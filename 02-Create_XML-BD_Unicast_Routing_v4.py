########################
# Add bds to the AEPs #
########################


# Excel File - for jinja Template
# Column 0  - Group
# Column 02 - Tenant - bd['tenant_name']
# Column 04 - BD - bd['bd_name']
# Column 03 - VRF - bd['vrf_ctx_name']
# Column 07 - L3out - bd['l3out_nameTo_out']
# Column 05 - Subnet - bd['ip_subnet']
# Column 06 - Unicast Routing - bd['unicast_routing']

from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader
#from pprint import pprint

# Load worksheet
wb = load_workbook('Input Files/NXOS_ACI_Info_baseline.xlsx')
sheet = wb["BD"]


# Create a list to get datas
infoList = []
tenants = []
bds = []

# Itere sobre as linhas na planilha
for row in list(sheet.iter_rows(values_only=True))[1:]:
    
    tenant = { 'name': row[2] } 
    bd = { 'tenant_name': row[2],'bd_name': row[4],'vrf_ctx_name': row[3], 'l3out_nameTo_out': row[7], 'ip_subnet': row[5], 'unicast_routing': row[6] }
    
    # Adiciono na lista
    bds.append(bd)
    tenants.append(tenant)

    line = {
            'tenants': tenants,
            'group' : row[0],
            'bds': bds
     }

infoList.append(line)

for client in infoList:
    #context = {'client': client} 
    # Load the template Jinja
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template('templates/TEMPLATE_BD.xml.jinja')

    # Fill teamplate with the datas
    output = template.render(client)

    # path file output
    path_file = f"Output/Grp_{client['group']}-BD_Unicast_Routing.xml"
    
    # Save the file output
    with open(path_file, 'w') as file:
        file.write(output)
        print("{:0<4} {} {:0<7}".format("#########", path_file.removeprefix("Output/"), "#########"))
    
    ########################################
    # Save the file output rollback
    path_file_rollback = f"Output/Grp_{client['group']}-BD_Unicast_Routing_Rollback.xml"

    client['rollback'] =  True

     # Fill teamplate with the datas
    output = template.render(client)

    # Save the File output rollback
    with open(path_file_rollback, 'w') as file_rollback:
        file_rollback.write(output)
        print("{:0<4} {} {:0<7}".format("#########",path_file_rollback.removeprefix("Output/"), "#########"))

    ########################################
    # Save the file output test_bd
    path_file_rollback = f"Output/1.1-Test_Scenario-BD_Unicast_Routing.xml"

    client['rollback'] =  False
    client['test_bd'] =  True

     # Fill teamplate with the datas
    output = template.render(client)

    # Save the File output rollback
    with open(path_file_rollback, 'w') as file_rollback:
        file_rollback.write(output)
        print("{:0<4} {} {:0<7}".format("#########",path_file_rollback.removeprefix("Output/"), "#########"))

    ########################################
     # Save the file output test_bd
    path_file_rollback = f"Output/1.1-Test_Scenario-BD_Unicast_Routing_Rollback.xml"

    client['test_bd'] =  False
    client['rollback_test_bd'] =  True

     # Fill teamplate with the datas
    output = template.render(client)

    # Save the File output rollback
    with open(path_file_rollback, 'w') as file_rollback:
        file_rollback.write(output)
        print("{:0<4} {} {:0<7}".format("#########",path_file_rollback.removeprefix("Output/"), "#########"))