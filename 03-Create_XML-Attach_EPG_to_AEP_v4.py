########################
# Add EPGs to the AEPs #
########################


# Excel File - for jinja Template
# Column 0 - Group
# Column 10 - AAEP - AAEPname
# Column 2 - Tenant - epg['tenant']
# Column 08 - EPG - epg['name']
# Column 09 - Application Profile - epg['app-profile']
# Column 1 - VLAN - epg['vlanId']

from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

# Load worksheet
wb = load_workbook('Input Files/NXOS_ACI_Info_baseline.xlsx')
sheet = wb["BD"]


# Create a list to get datas
infoList = []
AAEPs = []
epgs = []

# Itere sobre as linhas na planilha
for row in list(sheet.iter_rows(values_only=True))[1:]:
    
    epg = { 'vlanId': row[1],'name': row[8],'app-profile': row[9], 'tenant': row[2] }
    # Adiciono na lista
    epgs.append(epg)

    line = {
            'group' : row[0],
            'AAEPname' : row[10],
            'epgs': epgs
     }

infoList.append(line)

for client in infoList:
    #context = {'client': client} 
    # Load the template Jinja
    env = Environment(loader=FileSystemLoader('.'))
    template = env.get_template('templates/TEMPLATE_Attach_EPG_to_AEP.xml.jinja')

    # Fill teamplate with the datas
    output = template.render(client)
    
    # path file output
    path_file = f"Output/Attach_EPG_to_{client['AAEPname']}_all.xml"
    
    # Save the file output
    with open(path_file, 'w') as file:
        file.write(output)
        print("{:0<4} {} {:0<7}".format("#########", path_file.removeprefix("Output/"), "#########"))
    
    
    path_file_rollback = f"Output/Attach_EPG_to_{client['AAEPname']}_all_Rollback.xml"

    client['rollback'] =  True

     # Fill teamplate with the datas
    output = template.render(client)

    # Save the File output rollback
    with open(path_file_rollback, 'w') as file_rollback:
        file_rollback.write(output)
        print("{:0<4} {} {:0<7}".format("#########",path_file_rollback.removeprefix("Output/"), "#########"))

