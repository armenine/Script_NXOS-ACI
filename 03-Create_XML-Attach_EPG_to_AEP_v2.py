########################
# Add EPGs to the AEPs #
########################


# Excel File - Input_Data.xlsx
# Column 1 - Tenant
# Column 2 - VRF
# Column 03 - BD
# Column 04 - BD IP
# Column 05 - Unicast Routing
# Column 06 - L3Out association 
# Column 07 - EPG
# Column 08 - Application Profile
# Column 09 - AAEP
# Column 10 - VLAN    
# Column 11 - Preferred Group
# Column 12 - Description


import os
from openpyxl import load_workbook

local_path = os.getcwd()
excel_file = local_path+"\\Input Files\\NXOS_ACI_Info_baseline.xlsx"

vlan_order = {}
vlan_position = {}
tenant = {}
vrf = {}
bd_name = {}
bd_ip = {}
unicast_routing = {}
l3out = {}
epg_name = {}
app_profile = {}
aaep = {}
pref_grp = {}
description = {}

try:
    workbook = load_workbook(excel_file)
    workbook_new = False
except Exception as e:
    print("Workbook not found" + str(e))
    exit()

try:
    sheet = workbook["BD"]
    cols = sheet.max_column
    max_row = sheet.max_row
    print("################ \t Worksheet OK \t \t \t \t ################")

except Exception as e:
    print("GET Sheet Exception : ",e)
    exit()   

def write_xml_file(file, count, aaep, vlan, tenant, app_profile, epg_name):

    vlan = str(vlan)
    if count == 2:   
        file.write('<imdata totalCount="1"> \
                <infraAttEntityP dn="uni/infra/attentp-'+aaep+'" name="'+aaep+'" > \
            <infraGeneric name="default" > \
        <infraRsFuncToEpg encap="vlan-'+vlan+'" instrImedcy="lazy" mode="regular" primaryEncap="unknown" tDn="uni/tn-'+tenant+'/ap-'+app_profile+'/epg-'+epg_name+'" />')

    else:
      file.write('<infraRsFuncToEpg encap="vlan-'+vlan+'" instrImedcy="lazy" mode="regular" primaryEncap="unknown" tDn="uni/tn-'+tenant+'/ap-'+app_profile+'/epg-'+epg_name+'" />')
    
def write_xml_file_rollback(file, count, aaep, tenant, app_profile, epg_name):
    if count == 2:
        file.write('<imdata totalCount="1"> \
                <infraAttEntityP dn="uni/infra/attentp-'+aaep+'" name="'+aaep+'" > \
            <infraGeneric name="default" > \
        <infraRsFuncToEpg tDn="uni/tn-'+tenant+'/ap-'+app_profile+'/epg-'+epg_name+'" status="deleted"/>')

    else:
      file.write(' <infraRsFuncToEpg tDn="uni/tn-'+tenant+'/ap-'+app_profile+'/epg-'+epg_name+'" status="deleted"/>')

#Get Vlan Information for BDs
for i in range(2, max_row+1):
    #Vlan Order
    vlan_order[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=1).value
    
    #Vlan Position
    vlan_position[i] = sheet.cell(row=i, column=2).value
    
    #tenant
    tenant[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=3).value
    
    #vrf
    vrf[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=4).value
    
    #bd
    bd_name[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=5).value
    
    #IP Address
    bd_ip[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=6).value
    
    #Unicast Routing
    unicast_routing[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=7).value
    
    #L3Out
    l3out[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=8).value
    
    #EPG Name
    epg_name[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=9).value
    
    #Application Profile
    app_profile[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=10).value
    
    #AAEP
    aaep[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=11).value
    
    #Preferred Group
    pref_grp[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=12).value
    
    #Description
    description[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=13).value
    

#Bild XMLs
last_grp = None

for i in range(2, max_row+1):
    
    vlan = vlan_position[i]
    
    if last_grp is None:
        print("################ \t Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ ".xml \t \t ################")
        file_name = "\\Output\\Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ ".xml"
        file = open(local_path+file_name, 'w')
        write_xml_file(file, i, aaep[vlan], vlan, tenant[vlan], app_profile[vlan], epg_name[vlan])
        
        print("################ \t Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ "_Rollback.xml \t ################")
        file_name_rollback = "\\Output\\Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ "_Rollback.xml"
        file_rollback = open(local_path+file_name_rollback, 'w')
        write_xml_file_rollback(file_rollback, i, aaep[vlan],tenant[vlan], app_profile[vlan], epg_name[vlan])
        
        last_grp = vlan_order[vlan]
        
    elif last_grp != vlan_order[vlan]:
        file.write('</infraGeneric> \
                </infraAttEntityP> \
            </imdata>')
        print("################ \t Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ ".xml \t \t ################")
        file_name = "\\Output\\Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ ".xml"
        file = open(local_path+file_name, 'w')
        file.write('<imdata totalCount="1"> \
                <infraAttEntityP dn="uni/infra/attentp-'+aaep[vlan]+'" name="'+aaep[vlan]+'" > \
            <infraGeneric name="default" > ')
            
        write_xml_file(file, i, aaep[vlan], vlan, tenant[vlan], app_profile[vlan], epg_name[vlan])
        
        
        file_rollback.write('</infraGeneric> \
                </infraAttEntityP> \
            </imdata>')

        print("################ \t Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ "_Rollback.xml \t ################")
        file_name_rollback = "\\Output\\Grp_" + str(vlan_order[vlan])+"-Attach_EPG_to_" +aaep[vlan]+ "_Rollback.xml"
        file_rollback = open(local_path+file_name_rollback, 'w')
        file_rollback.write('<imdata totalCount="1"> \
                        <infraAttEntityP dn="uni/infra/attentp-'+aaep[vlan]+'" name="'+aaep[vlan]+'" > \
                    <infraGeneric name="default" > ')
        write_xml_file_rollback(file_rollback, i, aaep[vlan], tenant[vlan], app_profile[vlan], epg_name[vlan])
        
        last_grp = vlan_order[vlan]
    else:
        write_xml_file(file, i, aaep[vlan], vlan, tenant[vlan], app_profile[vlan], epg_name[vlan])
        write_xml_file_rollback(file_rollback, i, aaep[vlan], tenant[vlan], app_profile[vlan], epg_name[vlan])
    
    if i == max_row:
        file.write('</infraGeneric> \
                </infraAttEntityP> \
            </imdata>')
        
        
        file_rollback.write('</infraGeneric> \
                </infraAttEntityP> \
            </imdata>')
