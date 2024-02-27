
########################################
# Create BD and/or its Unicast Routing #
########################################


# Excel File - Input_Data.xlsx
# Column 01 - Group
# Column 02 - VLAN
# Column 03 - Tenant
# Column 04 - VRF
# Column 05 - BD
# Column 06 - BD IP
# Column 07 - Unicast Routing
# Column 08 - L3Out association 
# Column 09 - EPG
# Column 10 - AP
# Column 11 - AAEP
# Column 12 - Preferred Group
# Column 13 - Description


import os
from openpyxl import load_workbook

local_path = os.getcwd()
excel_file = local_path+"\\Input Files\\NXOS_ACI_Info_baseline.xlsx"

vlan_order = {}
vlan_position = {}
tenant = {}
vrf = {}
bd_name = {}
bd_ip = {}
unicast_routing = {}
l3out = {}
epg_name = {}
app_profile = {}
aaep = {}
pref_grp = {}
description = {}

try:
    workbook = load_workbook(excel_file)
    workbook_new = False
except Exception as e:
    print("Workbook not found" + str(e))
    exit()

try:
    sheet = workbook["BD"]
    cols = sheet.max_column
    max_row = sheet.max_row
    print("################ \t Worksheet OK \t \t \t \t ################")

except Exception as e:
    print("GET Sheet Exception : ",e)
    exit()   
    
#Write BD info into XML
def write_xml_file(file, count, tenant, description, bridge_domain, unicast_routing, l3out, vrf, bd_ip):
    if not description:
        description = ""
    if not l3out:
        l3out = ""
    if not bd_ip:
        bd_ip
        
    # Frist time need to add extra tag to XML #
    if count == 2:
    
        file.write('<fvTenant dn="uni/tn-'+tenant+'" name="'+tenant+'"> \
                        <fvBD OptimizeWanBandwidth="no" arpFlood="yes" descr="'+description+'"') 
    else:
        file.write('<fvBD OptimizeWanBandwidth="no" arpFlood="yes" descr="'+description+'"') 
    
    file.write(' dn="uni/tn-'+tenant+'/BD-'+bridge_domain+'" epClear="no" hostBasedRouting="no" intersiteBumTrafficAllow="no" intersiteL2Stretch="no" ipLearning="yes" ipv6McastAllow="no" limitIpLearnToSubnets="yes" mcastARPDrop="yes" mcastAllow="no" \
                            multiDstPktAct="bd-flood" name="'+bridge_domain+'" type="regular" unicastRoute="'+unicast_routing+'" unkMacUcastAct="flood" unkMcastAct="flood" v6unkMcastAct="flood" vmac="not-applicable">')
        
    if l3out and bd_ip:
        file.write('<fvRsBDToOut tnL3extOutName="'+l3out+'" />')

    file.write('<fvRsBdToEpRet resolveAct="resolve" /> \
                    <fvRsCtx tnFvCtxName="'+vrf+'" /> ')
    if bd_ip:
        file.write('<fvSubnet ip="'+bd_ip+'" ipDPLearning="enabled" scope="public,shared" virtual="no"/> \
                </fvBD>' )
    else:
        file.write('</fvBD>' )
        
#Write Rollback BD info into XML
def write_xml_file_rollback(file, count, tenant, description, bridge_domain, l3out, vrf, bd_ip):
    l3out_status = True
    bd_ip_status = True
    if not description:
        description = ""
    if not l3out:
        l3out_status = False
    if not bd_ip:
        bd_ip_status = False
        
    
              
    # Frist time need to add extra tag to XML #
    if count == 2:
    
        file.write('<fvTenant dn="uni/tn-'+tenant+'" name="'+tenant+'"> \
                        <fvBD OptimizeWanBandwidth="no" arpFlood="yes" descr="'+description+'"') 
    else:
        file.write('<fvBD OptimizeWanBandwidth="no" arpFlood="yes" descr="'+description+'"') 
    
    file.write(' dn="uni/tn-'+tenant+'/BD-'+bridge_domain+'" epClear="no" hostBasedRouting="no" intersiteBumTrafficAllow="no" intersiteL2Stretch="no" ipLearning="yes" ipv6McastAllow="no" limitIpLearnToSubnets="yes" mcastARPDrop="yes" mcastAllow="no" \
                            multiDstPktAct="bd-flood" name="'+bridge_domain+'" type="regular" unicastRoute="no" unkMacUcastAct="flood" unkMcastAct="flood" v6unkMcastAct="flood" vmac="not-applicable">')
        
    if l3out_status:
        file.write('<fvRsBDToOut tnL3extOutName="'+l3out+'" status="deleted" />')

    file.write('<fvRsBdToEpRet resolveAct="resolve" /> \
                    <fvRsCtx tnFvCtxName="'+vrf+'" /> ')
    if bd_ip_status:
        file.write('<fvSubnet ip="'+bd_ip+'" ipDPLearning="enabled" scope="public,shared" virtual="no" status="deleted"/> \
                </fvBD>' )
    else:
        file.write('</fvBD>' )

    
#Get Vlan Information for BDs
for i in range(2, max_row+1):
    #Vlan Order
    vlan_order[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=1).value
    
    #Vlan Position
    vlan_position[i] = sheet.cell(row=i, column=2).value
    
    #tenant
    tenant[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=3).value
    
    #vrf
    vrf[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=4).value
    
    #bd
    bd_name[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=5).value
    
    #IP Address
    bd_ip[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=6).value
    
    #Unicast Routing
    unicast_routing[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=7).value
    
    #L3Out
    l3out[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=8).value
    
    #EPG Name
    epg_name[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=9).value
    
    #Application Profile
    app_profile[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=10).value
    
    #AAEP
    aaep[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=11).value
    
    #Preferred Group
    pref_grp[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=12).value
    
    #Description
    description[sheet.cell(row=i, column=2).value] = sheet.cell(row=i, column=13).value
    
#Bild XMLs
last_grp = None

for i in range(2, max_row+1):
    
    vlan = vlan_position[i]
    
    if last_grp is None:
        print("################ \t Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing.xml \t \t ################")
        file_name = "\\Output\\Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing.xml"
        file = open(local_path+file_name, 'w')
        write_xml_file(file, i, tenant[vlan], description[vlan], bd_name[vlan], unicast_routing[vlan], l3out[vlan], vrf[vlan], bd_ip[vlan])
        
        print("################ \t Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing_Rollback.xml \t ################")
        file_name_rollback = "\\Output\\Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing_Rollback.xml"
        file_rollback = open(local_path+file_name_rollback, 'w')
        write_xml_file_rollback(file_rollback, i, tenant[vlan], description[vlan], bd_name[vlan], l3out[vlan], vrf[vlan], bd_ip[vlan])
        
        last_grp = vlan_order[vlan]
        
    elif last_grp != vlan_order[vlan]:
        file.write('</fvTenant>')
        print("################ \t Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing.xml \t \t ################")
        file_name = "\\Output\\Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing.xml"
        file = open(local_path+file_name, 'w')
        file.write('<fvTenant dn="uni/tn-'+tenant[vlan]+'" name="'+tenant[vlan]+'">')
            
        write_xml_file(file, i, tenant[vlan], description[vlan], bd_name[vlan], unicast_routing[vlan], l3out[vlan], vrf[vlan], bd_ip[vlan])
        
        
        file_rollback.write('</fvTenant>')
        print("################ \t Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing_Rollback.xml \t ################")
        file_name_rollback = "\\Output\\Grp_" + str(vlan_order[vlan])+"-BD_Unicast_Routing_Rollback.xml"
        file_rollback = open(local_path+file_name_rollback, 'w')
        file_rollback.write('<fvTenant dn="uni/tn-'+tenant[vlan]+'" name="'+tenant[vlan]+'">')
        write_xml_file_rollback(file_rollback, i, tenant[vlan], description[vlan], bd_name[vlan], l3out[vlan], vrf[vlan], bd_ip[vlan])
        
        last_grp = vlan_order[vlan]
    else:
        write_xml_file(file, i, tenant[vlan], description[vlan], bd_name[vlan], unicast_routing[vlan], l3out[vlan], vrf[vlan], bd_ip[vlan])
        write_xml_file_rollback(file_rollback, i, tenant[vlan], description[vlan], bd_name[vlan], l3out[vlan], vrf[vlan], bd_ip[vlan])
    
    if i == max_row:
        file.write('</fvTenant>')
        file_rollback.write('</fvTenant>')
    
    
    
    
    
    