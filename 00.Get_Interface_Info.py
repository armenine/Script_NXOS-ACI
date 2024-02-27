import os
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


hostname_regex = re.compile(r'^hostname')
int_regex = re.compile(r'^interface Vlan')
hsrp_ip_regex = re.compile (r'^    ip ')
ip_address_regex = re.compile (r'^  ip address ')
hsrp_group_regex = re.compile (r'^  hsrp [0-9]')
description_regex = re.compile(r'^  description ')
vrf_regex = re.compile(r'^  vrf member ')
ospf_regex = re.compile(r'^  ip router ospf ')
line_regex = re.compile(r'^\n')

file_order = 1

hostname_1 = ''
interface_1 = []
ip_address_1 = []
hsrp_ip_address_1 = []
hsrp_group_1 = []
description_1 = []
vrf_1 = []
ospf_1 = []

hostname_2 = ''
interface_2 = []
ip_address_2 = []
hsrp_ip_address_2 = []
hsrp_group_2 = []
description_2 = []
vrf_2 = []
ospf_2 = []

hostname_3 = ''
interface_3 = []
ip_address_3 = []
hsrp_ip_address_3 = []
hsrp_group_3 = []
description_3 = []
vrf_3 = []
ospf_3 = []

hostname_4 = ''
interface_4 = []
ip_address_4 = []
hsrp_ip_address_4 = []
hsrp_group_4 = []
description_4 = []
vrf_4 = []
ospf_4 = []


local_path = os.getcwd()
input_files_path = local_path + "\\Input Files\\"
files_list = os.listdir(input_files_path)

excel_file = input_files_path+"NXOS_ACI_Info_baseline.xlsx"


# Fuction to Create excel file with two sheet NXOS and BD
def create_excel():
    workbook = Workbook()
    workbook.save(excel_file)
    workbook = load_workbook(excel_file)
    workbook.create_sheet(index=1, title="NXOS")
    workbook.create_sheet(index=2, title="BD")
    return workbook

# Fuction to template for BD sheet with correct columns that need it.
def template_sheet_bd(wb):
    wb["BD"].cell(row=1, column=1, value="Group")
    wb["BD"].cell(row=1, column=2, value="VLAN")
    wb["BD"].cell(row=1, column=3, value="TENANT")
    wb["BD"].cell(row=1, column=4, value="VRF ACI")
    wb["BD"].cell(row=1, column=5, value="BD")
    wb["BD"].cell(row=1, column=6, value="IP")
    wb["BD"].cell(row=1, column=7, value="Unicast Routing")
    wb["BD"].cell(row=1, column=8, value="L3Out")
    wb["BD"].cell(row=1, column=9, value="EPG")
    wb["BD"].cell(row=1, column=10, value="Application Profile")
    wb["BD"].cell(row=1, column=11, value="AAEP")
    wb["BD"].cell(row=1, column=12, value="Preferred Group")
    wb["BD"].cell(row=1, column=13, value="Domain")
    wb["BD"].cell(row=1, column=14, value="Description")
    wb["BD"].cell(row=1, column=15, value="OSPF")
    wb["BD"].cell(row=1, column=16, value="VRF NX OS")
    wb["BD"].cell(row=1, column=17, value="HSRP Group")


# Condition If It's not exist excel file then create a excel file
if os.path.isfile(excel_file) is False:
    workbook = create_excel()
    del workbook['Sheet']
    template_sheet_bd(workbook)
    
# If exist excel file then verify some columns 
else:
    workbook = load_workbook(excel_file)
    for row in workbook["BD"].iter_rows(min_row=1, max_row=1, min_col=1, max_col=17, values_only=True):
        if "Group" in row and "EPG" in row and "HSRP Group" in row and "VLAN" in row:
            pass
        else:
            template_sheet_bd(workbook)

try:
    sheet = workbook["NXOS"]
    cols = sheet.max_column
    max_row = sheet.max_row
    print("################ Worksheet OK ################")
except Exception as e:
    print("GET Sheet Exception : ",e)
    exit()   


def search_in_files(inner_file):
    fp = open(input_files_path+inner_file, 'r')
    file = fp.readlines()
    hostname = ''
    interface = [] 
    ip_address = []
    ip_mask = ''
    hsrp_ip_address = []
    hsrp_group = []
    description = []
    vrf = []
    ospf = []    
    int = False
    control = 0
    
    for line in file:
        if hostname_regex.search(line):
            hostname = line.replace('hostname ','').strip()
                
        elif int_regex.search(line):
           interface.append(line.replace('interface Vlan','').strip())
           int = True
            
        elif ip_address_regex.search(line):
            ip_address.append(line.replace('  ip address ','').strip())
            ip_mask = line.strip()

        elif description_regex.search(line):
            description.append(line.replace('  description ','').strip())

        elif vrf_regex.search(line):
            vrf.append(line.replace('  vrf member ','').strip())

        elif ospf_regex.search(line):
            ospf.append(line.replace('ip router ','').strip())

        elif hsrp_group_regex.search(line):
            hsrp_group.append(line.replace('  hsrp ','').strip())
                
        elif hsrp_ip_regex.search(line):
            hsrp_ip_address.append((line.replace('    ip ','')).strip() + (ip_mask[ip_mask.find('/'):len(ip_mask)]).strip())
        
        elif line_regex.search(line) and int:
            int = False
            
            if len(ip_address) != len (interface):
                ip_address.append(" ")
                
            if len(description) != len (interface):
                description.append(" ")
            
            if len(vrf) != len (interface):
                vrf.append(" ")
                
            if len(ospf) != len (interface):
                ospf.append(" ")
                
            if len(hsrp_group) != len (interface):
                hsrp_group.append(" ")
                
            if len(hsrp_ip_address) != len (interface):
                hsrp_ip_address.append(" ")
            
            control = control + 1  
            
    return hostname, interface, ip_address, hsrp_ip_address, hsrp_group, description, vrf, ospf

def color_cell(inner_row):
    sheet.cell(row=inner_row, column=2).font = Font(color="FFFFFF", bold=True)
    sheet.cell(row=inner_row, column=2).fill = PatternFill(start_color="66B2FF", end_color="66B2FF", fill_type = "solid")
    
    sheet.cell(row=inner_row, column=3).font = Font(color="FFFFFF", bold=True)
    sheet.cell(row=inner_row, column=3).fill = PatternFill(start_color="66B2FF", end_color="66B2FF", fill_type = "solid")
    
    sheet.cell(row=inner_row, column=4).font = Font(color="FFFFFF", bold=True)
    sheet.cell(row=inner_row, column=4).fill = PatternFill(start_color="66B2FF", end_color="66B2FF", fill_type = "solid")
    
    sheet.cell(row=inner_row, column=5).font = Font(color="FFFFFF", bold=True)
    sheet.cell(row=inner_row, column=5).fill = PatternFill(start_color="66B2FF", end_color="66B2FF", fill_type = "solid")
    
    

for file in files_list:
    
    if 'txt' in file:
               
        if file_order == 1:
            hostname_1, interface_1, ip_address_1, hsrp_ip_address_1, hsrp_group_1, description_1, vrf_1, ospf_1 = search_in_files(file)
            
        if file_order == 2:
            hostname_2, interface_2, ip_address_2, hsrp_ip_address_2, hsrp_group_2, description_2, vrf_2, ospf_2 = search_in_files(file)
            
        if file_order == 3:
            hostname_3, interface_3, ip_address_3, hsrp_ip_address_3, hsrp_group_3, description_3, vrf_3, ospf_3 = search_in_files(file)
            
        if file_order == 4:
            hostname_4, interface_4, ip_address_4, hsrp_ip_address_4, hsrp_group_4, description_4, vrf_4, ospf_4 = search_in_files(file)
                        
        file_order = file_order + 1
        

row_excel = 2
next_row = False


sheet.cell(row=1, column=2, value= hostname_1)
sheet.cell(row=1, column=3, value= hostname_2)
sheet.cell(row=1, column=4, value= hostname_3)
sheet.cell(row=1, column=5, value= hostname_4)

index_vlan = 0

for i in range (1, 4000):
    if str(i) in interface_1:
        sheet.cell(row=row_excel, column=2, value= i)
        color_cell(row_excel)
        
        next_row = True
        index_vlan = interface_1.index(str(i))
        sheet.cell(row=row_excel+1, column=2, value= ip_address_1[index_vlan])
        sheet.cell(row=row_excel+2, column=2, value= hsrp_ip_address_1[index_vlan])
        sheet.cell(row=row_excel+3, column=2, value= hsrp_group_1[index_vlan])
        sheet.cell(row=row_excel+4, column=2, value= description_1[index_vlan])
        sheet.cell(row=row_excel+5, column=2, value= vrf_1[index_vlan])
        sheet.cell(row=row_excel+6, column=2, value= ospf_1[index_vlan])
        
        
    if str(i) in interface_2:
        sheet.cell(row=row_excel, column=3, value= i)
        color_cell(row_excel)
        
        next_row = True
        index_vlan = interface_2.index(str(i))
        sheet.cell(row=row_excel+1, column=3, value= ip_address_2[index_vlan])
        sheet.cell(row=row_excel+2, column=3, value= hsrp_ip_address_2[index_vlan])
        sheet.cell(row=row_excel+3, column=3, value= hsrp_group_2[index_vlan])
        sheet.cell(row=row_excel+4, column=3, value= description_2[index_vlan])
        sheet.cell(row=row_excel+5, column=3, value= vrf_2[index_vlan])
        sheet.cell(row=row_excel+6, column=3, value= ospf_2[index_vlan])
        
    if str(i) in interface_3:
        sheet.cell(row=row_excel, column=4, value= i)
        color_cell(row_excel)        

        next_row = True
        index_vlan = interface_3.index(str(i))
        sheet.cell(row=row_excel+1, column=4, value= ip_address_3[index_vlan])
        sheet.cell(row=row_excel+2, column=4, value= hsrp_ip_address_3[index_vlan])
        sheet.cell(row=row_excel+3, column=4, value= hsrp_group_3[index_vlan])
        sheet.cell(row=row_excel+4, column=4, value= description_3[index_vlan])
        sheet.cell(row=row_excel+5, column=4, value= vrf_3[index_vlan])
        sheet.cell(row=row_excel+6, column=4, value= ospf_3[index_vlan])
        
    if str(i) in interface_4:
        sheet.cell(row=row_excel, column=5, value= i)
        color_cell(row_excel)
        
        next_row = True
        index_vlan = interface_4.index(str(i))
        sheet.cell(row=row_excel+1, column=5, value= ip_address_4[index_vlan])
        sheet.cell(row=row_excel+2, column=5, value= hsrp_ip_address_4[index_vlan])
        sheet.cell(row=row_excel+3, column=5, value= hsrp_group_4[index_vlan])
        sheet.cell(row=row_excel+4, column=5, value= description_4[index_vlan])
        sheet.cell(row=row_excel+5, column=5, value= vrf_4[index_vlan])
        sheet.cell(row=row_excel+6, column=5, value= ospf_4[index_vlan])
        
    if next_row:
        sheet.cell(row=row_excel, column=1, value= "Interface Vlan")
        sheet.cell(row=row_excel, column=1).font = Font(color="FFFFFF", bold=True)
        sheet.cell(row=row_excel, column=1).fill = PatternFill(start_color="66B2FF", end_color="66B2FF", fill_type = "solid")
        
        sheet.cell(row=row_excel+1, column=1, value= "IP Address")
        sheet.cell(row=row_excel+2, column=1, value= "HSRP")
        sheet.cell(row=row_excel+3, column=1, value= "HSRP Group")
        sheet.cell(row=row_excel+4, column=1, value= "Description")
        sheet.cell(row=row_excel+5, column=1, value= "VRF")
        sheet.cell(row=row_excel+6, column=1, value= "OSPF")
        row_excel = row_excel + 8
        next_row = False
                
workbook.save(excel_file)
print("Workbook NXOS_ACI_Info.xlsx saved")