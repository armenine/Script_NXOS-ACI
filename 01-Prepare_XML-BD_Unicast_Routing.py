import os
from openpyxl import load_workbook

vlans_selection = []
vlans_order = []

cel_interface_1 = {}
cel_interface_2 = {}
cel_interface_3 = {}
cel_interface_4 = {}

local_path = os.getcwd()

excel_file = local_path+"\\Input Files\\NXOS_ACI_Info_baseline.xlsx"

try:
    workbook = load_workbook(excel_file)
    
except Exception as e:
    print("Workbook not found" + str(e))
    exit()

try:
    sheet_nxos = workbook["NXOS"]
    max_cols_nxos = sheet_nxos.max_column
    max_row_nxos = sheet_nxos.max_row
    
    sheet_bd = workbook["BD"]
    max_cols_bd = sheet_bd.max_column
    max_row_bd = sheet_bd.max_row
    
    print("################ \t Worksheet OK \t \t \t \t ################")
    
    #print ("max_cols_nxos ",max_cols_nxos)
    #print("max_row_nxos ", max_row_nxos)
    #
    #print ("max_cols_bd ",max_cols_bd)
    #print("max_row_bd ", max_row_bd)
    
except Exception as e:
    print("GET Sheet Exception : ",e)
    exit()   

def convert_vlan_name_to_epg_bd(inner_vlan, inner_row):
    
    #Adding EPG names based at VLANs number
    if len(str(inner_vlan)) == 1:
        sheet_bd.cell(row = inner_row, column= 5).value = 'BD_000' + str(inner_vlan)
        sheet_bd.cell(row = inner_row, column= 9).value = 'EPG_000' + str(inner_vlan)
    elif len(str(inner_vlan)) == 2:
        sheet_bd.cell(row = inner_row, column= 5).value = 'BD_00' + str(inner_vlan)
        sheet_bd.cell(row = inner_row, column= 9).value = 'EPG_00' + str(inner_vlan)
    elif len(str(inner_vlan)) == 3:
        sheet_bd.cell(row = inner_row, column= 5).value = 'BD_0' + str(inner_vlan)
        sheet_bd.cell(row = inner_row, column= 9).value = 'EPG_0' + str(inner_vlan)
    else:
        sheet_bd.cell(row = inner_row, column= 5).value = 'BD_' + str(inner_vlan)
        sheet_bd.cell(row = inner_row, column= 9).value = 'EPG_' + str(inner_vlan)
 
def add_info_excel(inner_row, ip_addr):
    #Adding HSRP IP as BD IP at excel
    sheet_bd.cell(row=inner_row+2, column=6, value = ip_addr)
    
    #Adding Unicast Routing
    sheet_bd.cell(row=inner_row+2, column=7, value = "yes")
    
    #Adding Preffered Group
    sheet_bd.cell(row=inner_row+2, column=12, value = "yes")
    
    #Adding EPG names based at VLANs number
    convert_vlan_name_to_epg_bd(vlan, i+2)
       

for i in range(2, max_row_bd+1):
    
    vlans_order.append(sheet_bd.cell(row=i,column=1).value)
    vlans_selection.append(sheet_bd.cell(row=i,column=2).value)

for i in range(2, max_row_nxos+1, 8):
    cel_interface_1[sheet_nxos.cell(row=i,column=2).value] = i
    cel_interface_2[sheet_nxos.cell(row=i,column=3).value] = i
    cel_interface_3[sheet_nxos.cell(row=i,column=4).value] = i
    cel_interface_4[sheet_nxos.cell(row=i,column=5).value] = i
    

#Look for Input VLANs at 4 NXOS infos
for i in range(0, len(vlans_selection)):
    vlan = vlans_selection[i]
    row_vlan_1 = cel_interface_1.get(vlan) 
    row_vlan_2 = cel_interface_2.get(vlan) 
    row_vlan_3 = cel_interface_3.get(vlan) 
    row_vlan_4 = cel_interface_4.get(vlan) 
      
    if vlans_selection[i] in cel_interface_1 and sheet_nxos.cell(row=row_vlan_1 +2,column=2).value is not None:
        add_info_excel(i,sheet_nxos.cell(row=row_vlan_1 +2,column=2).value)
    elif vlans_selection[i] in cel_interface_2 and sheet_nxos.cell(row=row_vlan_2 +2,column=3).value is not None:
        add_info_excel(i,sheet_nxos.cell(row=row_vlan_2 +2,column=3).value)
    elif vlans_selection[i] in cel_interface_3 and sheet_nxos.cell(row=row_vlan_3 +2,column=4).value is not None:
        add_info_excel(i,sheet_nxos.cell(row=row_vlan_3 +2,column=4).value)
    elif vlans_selection[i] in cel_interface_4 and sheet_nxos.cell(row=row_vlan_4 +2,column=5).value is not None:
        add_info_excel(i,sheet_nxos.cell(row=row_vlan_4 +2,column=5).value)
    else:
        convert_vlan_name_to_epg_bd(vlan, i+2)
        #Adding Unicast Routing
        sheet_bd.cell(row=i+2, column=7, value = "no")
        
        #Adding Preffered Group
        sheet_bd.cell(row=i+2, column=12, value = "yes")



workbook.save(excel_file)
print("################ \t Workbook NXOS_ACI_Info.xlsx saved \t ################")